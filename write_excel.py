from tkinter import *
import tkinter as tk
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib


main=Tk()
main.title("Login Form")
main.geometry("800x250")
main.config(highlightbackground="black",highlightthickness=2)

        
file = pathlib.Path("Backened_Data.xlsx")
if file.exists ():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Full Name"
    sheet["B1"]="Username"
    sheet["C1"]="Password"
    sheet["D1"]="Mail"
    
    file.save("Backened_Data.xlsx")


def show_Password():
    if var.get() ==1:
        passEntry.config(show='')
        
    if var.get() ==0:
        passEntry.config(show='*')

def submit():
    y=name.get()
    z=user.get()
    z1=passEntry.get()
    y1=emailentry.get()
    print(y)
    print(z)
    print(z1)
    print(y1)
    
    
    file=openpyxl.load_workbook("Backened_Data.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=y)
    sheet.cell(column=2,row=sheet.max_row,value=z)
    sheet.cell(column=3,row=sheet.max_row,value=z1)
    sheet.cell(column=4,row=sheet.max_row,value=y1)
       
    file.save("Backened_Data.xlsx")
'''
    xlfile = pd.read_excel('Backened_Data.xlsx', 'Sheet') # reading xl file 
    xlfile.to_csv('Backened_Data.csv', index=False)#conversion to csv
'''
  
frame1 = LabelFrame(main, text = 'Login Details:').pack(expand = 'yes', fill = 'both')

Label(frame1,text="Name:").place(x=50,y=30)
Label(frame1,text="Username:").place(x=50,y=70)
Label(frame1,text="Password:").place(x=50,y=110)
Label(main,text="Mail ID:").place(x=50,y=150)

name = Entry(frame1)
name.place(x=250,y=30)
user =Entry(frame1)
user.place(x=250,y=70)
password = StringVar() 
passEntry = Entry(frame1, textvariable=password, show='*')
passEntry.place(x=250,y=110)

emailentry = Entry(frame1)
emailentry.place(x=250,y=150,width=250)

var =IntVar()
var.set('0')
chkbutton = Checkbutton(frame1, text='Show',variable=var, onvalue=1, offvalue=0, \
                        command=show_Password).place(x=350,y=110)

Button(frame1,text = "Subscribe",command = submit).place(x=400,y=200)

main.mainloop()